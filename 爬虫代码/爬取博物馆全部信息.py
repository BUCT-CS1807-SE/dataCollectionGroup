import requests
from lxml import etree
import pandas as pd
import openpyxl
from tqdm import tqdm
import re

# 1.指定url
all_data = list(pd.read_csv('museum_info.csv')['博物馆对应页面的编号'])  # 拿到所有页面的编号,encoding='gbk'
all_id = list(pd.read_csv('museum_info.csv')['博物馆编号（网站当中给的排号）'])
all_name = list(pd.read_csv('museum_info.csv')['博物馆名称'])
all_type = list(pd.read_csv('museum_info.csv')['博物馆的类型（四个类型当中的一种）'])


def visit_site_basic_information():  # 开始按照网址访问网站，并填写基本信息
    #UA伪装
    headers = {
        'User-Agent' : 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.88 Safari/537.36'
    }
    #读取excel，用以知道对应的属性都在第几列
    fixed_title = list(pd.read_excel('博物馆信息.xlsx'))
    #读取文件：load_workbook
    wb = openpyxl.load_workbook('博物馆信息.xlsx')
    sheet = wb.active
    for i in tqdm(range(len(all_data))):
       #if i < 2 :
            url = 'https://www.maigoo.com/citiao/' + str(all_data[i]) + '.html'  # 指定网址
            print(url)
            
            #获取源码内容
            page_text = requests.get(url=url, headers=headers).text
            print(page_text)
            
            #获取周围的景点################################
            result = re.findall('class="color1">(.*?)<span class="font12">(.*?)</span></a>',page_text,re.S)
            #print(result)
            k = 0
            string=""
            length = len(result)
            while(k<length):
                if(k!= length-1):
                    k=k+1
                    string=string+result[k][0]+result[k][1]+","
                else:
                    string=string+result[k][0]+result[k][1]
                    break
            #print(string)
            sheet.cell(row=i + 2, column= 15).value = string
            
            #生成实例对象
            tree = etree.HTML(page_text)

            #获取地址信息，票价，开馆时间、建议游玩时间、博物馆类型、国际级别等信息###############################
            # 下面一次性可以将博物馆所处于的地址信息，票价，开馆时间、建议游玩时间、博物馆类型、国际级别、第几批选入的国家一级博物馆、旅游景区的级别
            basic_information_title = tree.xpath('//div[@class="lbox"]/em[@class="fcolor bdcolor"]/text()')
            basic_information = tree.xpath('//div[@class="lbox"]/span[@class="c666 dhidden"]/text()')

            #获取简介################################
            brief_introduction = tree.xpath('//div[@class="cont c666 font16"]/text()')
            print(brief_introduction)
            sheet.cell(row=i + 2, column= 12).value = brief_introduction[0]
            
            # 因为title的格式不符合csv当中的格式，因此需要进行修改
            if '官网：' in basic_information_title:
                basic_information_title.remove('官网：')
            print(basic_information_title)
            
            #将那个museum_info表的内容写进去
            sheet.cell(row=i + 2, column= 1).value = all_id[i]
            sheet.cell(row=i + 2, column= 2).value = all_name[i]
            sheet.cell(row=i + 2, column= 3).value = all_type[i]
            sheet.cell(row=i + 2, column= 11).value = all_data[i]
            
            for j in range(len(basic_information_title)):
                temp_title = str(basic_information_title[j])
                temp_title = temp_title[:-1]
                if temp_title in fixed_title:
                    #获取到对应的列
                    index = fixed_title.index(temp_title)
                    #将对应的行列的数据替换
                    sheet.cell(row=i + 2, column=index + 1).value = basic_information[j]
            print(basic_information)
    wb.save('博物馆信息.xlsx')


visit_site_basic_information()
